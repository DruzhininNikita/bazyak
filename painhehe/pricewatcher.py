import tkinter as tk
from tkinter import messagebox
import customtkinter as ctk
import threading
import time
import re
import os
import platform
import subprocess
import urllib.parse as urlparse
from datetime import datetime
from pathlib import Path
from tkinter import filedialog
from PIL import Image, ImageTk
from playwright.sync_api import sync_playwright
from plyer import notification


try:
    PLAYWRIGHT_OK = True
except ImportError:
    PLAYWRIGHT_OK = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

BASE_DIR = Path(__file__).parent / "PriceWatcher"
BASE_DIR.mkdir(exist_ok=True)
EXCEL_PATH = BASE_DIR / "price_history.xlsx"

# ─────────────────────────────────────────────────────────────────────────────
# ПАЛИТРЫ ТЕМ (base + accessibility)
# ─────────────────────────────────────────────────────────────────────────────
FONT_SIZES = {
    "default": {
        "title": 26, "subtitle": 12, "label": 13, "label_sm": 11,
        "input": 13, "input_price": 17, "btn": 13, "btn_big": 17,
        "log": 12, "stat_big": 24, "stat_label": 14,
        "tab": 14,
    },
    "large": {
        "title": 34, "subtitle": 16, "label": 17, "label_sm": 14,
        "input": 17, "input_price": 22, "btn": 17, "btn_big": 22,
        "log": 16, "stat_big": 31, "stat_label": 18,
        "tab": 18,
    },
}

def _sizes(base_key):
    s = FONT_SIZES[base_key]
    return s

PALETTES = {
    # ── Dark (Классический темный в стиле VS Code / Discord) ──────────────────
    "Dark": {
        "window_bg": "#1E1E1E", "header_bg": "#252526", "tab_bg": "#1E1E1E",
        "card_bg": "#2D2D30", "card_bg2": "#252526", "input_bg": "#3C3C3C",
        "log_bg": "#1E1E1E", "status_bg": "#007ACC",
        
        "accent": "#007ACC", "accent_hover": "#1E8AD4",
        
        "btn_add_bg": "#0E639C", "btn_add_hover": "#1177BB",
        "btn_start_bg": "#0E8C3A", "btn_start_hover": "#10A345",
        "btn_stop_bg": "#C72E0F", "btn_stop_hover": "#E03413",
        "btn_excel_bg": "#1D6E43", "btn_excel_hover": "#238636",
        "btn_paste_bg": "#3E3E42", "btn_paste_hover": "#505050",
        
        "label_main": "#F1F1F1", "label_sub": "#CCCCCC", "label_dim": "#858585",
        "log_text": "#CCCCCC", "status_ok": "#3FB950",
        "queue_url": "#58A6FF", "queue_price": "#3FB950",
        "section_label": "#007ACC",
        
        "tab_selector": "#007ACC", "tab_text": "#CCCCCC",
        "entry_text": "#F1F1F1",
    },
    
    # ── Light (Мягкая светлая тема с теплыми оттенками) ───────────────────────
    "Light": {
        "window_bg": "#FAFAFA", "header_bg": "#F5F5F5", "tab_bg": "#FAFAFA",
        "card_bg": "#FFFFFF", "card_bg2": "#F8F9FA", "input_bg": "#FFFFFF",
        "log_bg": "#FFFFFF", "status_bg": "#4A90E2",
        
        "accent": "#4A90E2", "accent_hover": "#357ABD",
        
        "btn_add_bg": "#5B9BD5", "btn_add_hover": "#4A8BC9",
        "btn_start_bg": "#6BBA6E", "btn_start_hover": "#5AA85E",
        "btn_stop_bg": "#E57373", "btn_stop_hover": "#D32F2F",
        "btn_excel_bg": "#6BBA6E", "btn_excel_hover": "#5AA85E",
        "btn_paste_bg": "#E0E0E0", "btn_paste_hover": "#D0D0D0",
        
        "label_main": "#2C3E50", "label_sub": "#546E7A", "label_dim": "#90A4AE",
        "log_text": "#37474F", "status_ok": "#43A047",
        "queue_url": "#1E88E5", "queue_price": "#43A047",
        "section_label": "#4A90E2",
        
        "tab_selector": "#4A90E2", "tab_text": "#FFFFFF",
        "entry_text": "#2C3E50",
    },
    
    # ── High Contrast (Настоящая высокая контрастность) ───────────────────────
    "HighContrast": {
        "window_bg": "#000000", "header_bg": "#0A0A0A", "tab_bg": "#000000",
        "card_bg": "#0A0A0A", "card_bg2": "#050505", "input_bg": "#000000",
        "log_bg": "#000000", "status_bg": "#000000",
        
        "accent": "#FFD700", "accent_hover": "#FFC107", 
        
        "btn_add_bg": "#000000", "btn_add_hover": "#FFD700",
        "btn_start_bg": "#00FF00", "btn_start_hover": "#00CC00",
        "btn_stop_bg": "#FF4444", "btn_stop_hover": "#CC0000",
        "btn_excel_bg": "#00BFFF", "btn_excel_hover": "#0099CC",
        "btn_paste_bg": "#000000", "btn_paste_hover": "#FFD700",
        
        "label_main": "#FFFFFF", "label_sub": "#FFD700", "label_dim": "#00BFFF",
        "log_text": "#00FF00", "status_ok": "#00FF00",
        "queue_url": "#00BFFF", "queue_price": "#FFD700",
        "section_label": "#FFD700",
        
        "tab_selector": "#FFD700", "tab_text": "#000000",
        "entry_text": "#FFFFFF"
    },
    
    # ── Colorblind (Оптимизировано для дальтоников) ───────────────────────────
    "Colorblind": {
        "window_bg": "#F5F5F5", "header_bg": "#EEEEEE", "tab_bg": "#F5F5F5",
        "card_bg": "#FFFFFF", "card_bg2": "#FAFAFA", "input_bg": "#FFFFFF",
        "log_bg": "#FFFFFF", "status_bg": "#000000",
        
        "accent": "#005A9C", "accent_hover": "#004578",
        
        "btn_add_bg": "#005A9C", "btn_add_hover": "#004578",
        "btn_start_bg": "#008837", "btn_start_hover": "#006B2C",
        "btn_stop_bg": "#CC3311", "btn_stop_hover": "#A62A0E",
        "btn_excel_bg": "#E69F00", "btn_excel_hover": "#D18F00",
        "btn_paste_bg": "#D9D9D9", "btn_paste_hover": "#C0C0C0",
        
        "label_main": "#1A1A1A", "label_sub": "#4D4D4D", "label_dim": "#737373",
        "log_text": "#1A1A1A", "status_ok": "#008837",
        "queue_url": "#005A9C", "queue_price": "#E69F00",
        "section_label": "#005A9C",
        
        "tab_selector": "#005A9C", "tab_text": "#FFFFFF",
        "entry_text": "#1A1A1A",
    },
}

ACCESSIBILITY_LABELS = {
    "Dark": "Стандартная тёмная",
    "Light": "Стандартная светлая",
    "HighContrast": "Высокая контрастность",
    "Colorblind": "Режим для дальтоников",
}

def send_os_notification(title, message):
    system = platform.system()
    try:
        if system == "Windows":
            try:
                notification.notify(title=title, message=message, app_name="PriceWatcher", timeout=8)
            except Exception:
                ps_cmd = (
                    '[Windows.UI.Notifications.ToastNotificationManager, Windows.UI.Notifications, ContentType = WindowsRuntime] > $null;'
                    '$template = [Windows.UI.Notifications.ToastTemplateType]::ToastText02;'
                    '$xml = [Windows.UI.Notifications.ToastNotificationManager]::GetTemplateContent($template);'
                    f'$xml.GetElementsByTagName("text")[0].AppendChild($xml.CreateTextNode("{title}")) > $null;'
                    f'$xml.GetElementsByTagName("text")[1].AppendChild($xml.CreateTextNode("{message}")) > $null;'
                    '$toast = [Windows.UI.Notifications.ToastNotification]::new($xml);'
                    '[Windows.UI.Notifications.ToastNotificationManager]::CreateToastNotifier("PriceWatcher").Show($toast);'
                )
                subprocess.run(["powershell", "-Command", ps_cmd], capture_output=True)
        elif system == "Linux":
            subprocess.run(["notify-send", "-a", "PriceWatcher", title, message], capture_output=True)
        elif system == "Darwin":
            subprocess.run(["osascript", "-e", f'display notification "{message}" with title "{title}"'], capture_output=True)
    except Exception:
        pass


def init_excel():
    if not OPENPYXL_OK or EXCEL_PATH.exists():
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "История цен"
    hf = PatternFill("solid", fgColor="FF00BF")
    ff = Font(bold=True, color="0755E7", size=11)
    thin = Side(style="thin", color="FFB399")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers = ["Дата/Время", "URL товара", "Название", "Цена (руб.)", "Цель (руб.)", "Статус", "Платформа"]
    col_widths = [20, 55, 35, 14, 14, 16, 12]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = ff; cell.fill = hf
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = brd
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws2 = wb.create_sheet("Статистика")
    ws2["A1"] = "Статистика мониторинга"
    ws2["A1"].font = Font(bold=True, size=13, color="E94560")
    ws2["A3"] = "Всего проверок:"; ws2["B3"] = "=COUNTA('История цен'!A2:A10000)"
    ws2["A4"] = "Найдено скидок:"; ws2["B4"] = "=COUNTIF('История цен'!F2:F10000,\"*Скидка*\")"
    ws2["A5"] = "Мин. цена:";       ws2["B5"] = "=IFERROR(MIN('История цен'!D2:D10000),\"-\")"
    ws2["A6"] = "Макс. цена:";      ws2["B6"] = "=IFERROR(MAX('История цен'!D2:D10000),\"-\")"
    for r in range(3, 7): ws2.cell(r, 1).font = Font(bold=True)
    wb.save(EXCEL_PATH)


def log_to_excel(url, name, price, target_price, status):
    if not OPENPYXL_OK: return
    try:
        init_excel()
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["История цен"]
        thin = Side(style="thin", color="444444")
        brd = Border(left=thin, right=thin, top=thin, bottom=thin)
        nr = ws.max_row + 1
        rf = PatternFill("solid", fgColor="F0E7EA" if nr % 2 == 0 else "FEFEFF")
        sc = "00AA00" if "Скидка" in status else "DDDDDD"
        row_data = [datetime.now().strftime("%d.%m.%Y %H:%M:%S"), url, name or "—",
                    price if isinstance(price, int) else "—", target_price, status, platform.system()]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=nr, column=col, value=val)
            cell.fill = rf; cell.border = brd
            cell.alignment = Alignment(vertical="center")
            if col == 6: cell.font = Font(bold=True, color=sc)
            if col == 4 and isinstance(val, int): cell.number_format = '#,##0 "₽"'
        wb.save(EXCEL_PATH)
    except Exception as e:
        print(f"Excel error: {e}")


PRICE_SELECTORS = [
    'span[data-auto="snippet-price-current"] b',
    'span[data-auto="snippet-price-current"]',
    '[data-auto="price-value"]',
    'div[data-zone-name="price"] span',
    '.Price_price__pkgD2',
    'h3[data-auto="snippet-price-current"]',
    '[class*="Price"] [class*="value"]',
    'span[class*="price"]',
]
NAME_SELECTORS = ['h1[data-auto="productName"]', '[data-auto="snippet-title"]', 'h1[class*="Title"]', 'h1']

def parse_price_from_page(page):
    name = None
    for sel in NAME_SELECTORS:
        try:
            el = page.locator(sel).first
            if el.is_visible(timeout=2000):
                name = el.inner_text().strip()[:80]; break
        except Exception: pass
    for sel in PRICE_SELECTORS:
        try:
            for el in page.locator(sel).all():
                try:
                    if el.is_visible(timeout=1000):
                        digits = re.sub(r'\D', '', el.inner_text().strip())
                        if digits and 100 <= int(digits) <= 10_000_000:
                            return int(digits), name
                except Exception: continue
        except Exception: continue
    try:
        matches = re.findall(r'"price"\s*:\s*"?(\d{3,7})"?', page.content())
        prices = [int(m) for m in matches if 100 <= int(m) <= 10_000_000]
        if prices: return min(prices), name
    except Exception: pass
    return None, name


def clean_yandex_url(url):
    url = url.strip()
    if not url.startswith("http"): url = "https://" + url
    try:
        parsed = urlparse.urlparse(url)
        qs = urlparse.parse_qs(parsed.query, keep_blank_values=False)
        allowed = {k: qs[k][0] for k in ("sku", "skuId", "modelid") if k in qs}
        return urlparse.urlunparse((parsed.scheme or "https", parsed.netloc or "market.yandex.ru",
                                    parsed.path, "", urlparse.urlencode(allowed), ""))
    except Exception: return url


INTERVAL_OPTIONS = {
    "5 минут": 300, "15 минут": 900, "30 минут": 1800,
    "1 час": 3600, "2 часа": 7200, "4 часа": 14400,
    "8 часов": 28800, "24 часа": 86400,
}


class PriceWatcherApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("PriceWatcher")
        self.geometry("660x840")
        self.minsize(560, 720)
        self.resizable(True, True)

        self._base_theme = "Dark"       # Dark / Light / HighContrast / Colorblind
        self._font_mode  = "default"   # default / large
        self._current_theme = "Dark"
        ctk.set_appearance_mode("Dark")
        ctk.set_default_color_theme("blue")

        self.is_monitoring = False
        self.monitor_thread = None
        self.monitored_items = []
        self._bg_image = None
        self.bg_path = None
        self._session = {"checks": 0, "found": 0, "deals": 0, "errors": 0}
        self._themed_widgets = []

        # Все виджеты, которые нужно перестраивать при смене размера шрифта
        self._rebuildable_widgets = []

        init_excel()
        self._build_ui()

    # ── Цвет текущей темы ─────────────────────────────────────────────────────
    def _t(self, key):
        pal = PALETTES.get(self._current_theme, PALETTES["Dark"])
        return pal.get(key, PALETTES["Dark"].get(key, "#FFFFFF"))

    # ── Размер шрифта ─────────────────────────────────────────────────────────
    def _fs(self, key):
        return _sizes(self._font_mode)[key]

    def _f(self, key, bold=False):
        sz = self._fs(key)
        return ("Segoe UI", sz, "bold") if bold else ("Segoe UI", sz)

    # ── Регистрация виджетов ──────────────────────────────────────────────────
    def user_signup(self, widget, role, attr="fg_color"):
        self._themed_widgets.append((widget, role, attr))
        return widget

    def _rebuildable(self, widget):
        self._rebuildable_widgets.append(widget)
        return widget

    # ── Применить тему (цвет) ─────────────────────────────────────────────────
    def _apply_theme(self):
        for (widget, role, attr) in self._themed_widgets:
            try:
                color = self._t(role)
                if attr == "fg_color":       widget.configure(fg_color=color)
                elif attr == "text_color":   widget.configure(text_color=color)
                elif attr == "border_color": widget.configure(border_color=color)
            except Exception: pass
        try: self.log_box.configure(fg_color=self._t("log_bg"), text_color=self._t("log_text"))
        except Exception: pass
        try: self.status_bar.configure(fg_color=self._t("status_bg"))
        except Exception: pass
        # Перестраиваем очередь и статистику (шрифты)
        self._refresh_queue_display()

    # ── Смена темы для слабовидящих ───────────────────────────────────────────
    def _set_accessibility_theme(self, theme_key):
        """Переключает палитру доступности."""
        self._base_theme = theme_key
        self._current_theme = theme_key
        # Исправленная строка:
        ctk.set_appearance_mode("Dark" if theme_key in ("Dark", "HighContrast") else "Light")
        self._apply_theme()
        self._rebuild_tabs()

    def _set_font_size(self, mode):
        """Переключает режим размера шрифта."""
        self._font_mode = mode
        self._rebuild_tabs()

    def _rebuild_tabs(self):
        """Полностью перестраивает вкладки при смене размера шрифта."""
        self._themed_widgets.clear()
        try:
            self.tabs.destroy()
        except Exception: pass
        self._build_tabs()
        self._apply_theme()

    # ── Построение UI ─────────────────────────────────────────────────────────
    def _build_ui(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self._build_header()
        self._build_tabs()
        self._build_statusbar()

    def _build_header(self):
        header = ctk.CTkFrame(self, fg_color=self._t("header_bg"), corner_radius=0, height=74)
        header.grid(row=0, column=0, sticky="ew")
        header.grid_columnconfigure(1, weight=1)
        header.grid_propagate(False)
        self._register(header, "header_bg")

        ctk.CTkLabel(header, text="⚡", font=("Segoe UI Emoji", self._fs("title"))).grid(

            row=0, column=0, rowspan=2, padx=18, pady=10)

        title_lbl = ctk.CTkLabel(header, text="PriceWatcher",
                                  font=("Georgia", self._fs("title"), "bold"),
                                  text_color=self._t("accent"))
        title_lbl.grid(row=0, column=1, sticky="sw", pady=(16, 0))
        self._register(title_lbl, "accent", "text_color")

        sub_lbl = ctk.CTkLabel(header, text="Яндекс.Маркет  •  Мониторинг цен",
                                font=("Segoe UI", self._fs("subtitle")),
                                text_color=self._t("label_dim"))
        sub_lbl.grid(row=1, column=1, sticky="nw", pady=(0, 12))
        self._register(sub_lbl, "label_dim", "text_color")

        settings_btn = ctk.CTkButton(
            header, text="⚙  Настройки", width=120, height=36,
            font=self._f("btn", bold=True),
            fg_color=self._t("btn_paste_bg"),
            hover_color=self._t("btn_paste_hover"),
            command=self._open_settings)
        settings_btn.grid(row=0, column=2, rowspan=2, padx=16)
        self._register(settings_btn, "btn_paste_bg")

    def _build_tabs(self):
        self.tabs = ctk.CTkTabview(self, fg_color=self._t("tab_bg"))
        self.tabs.grid(row=1, column=0, sticky="nsew", padx=0, pady=0)
        self._register(self.tabs, "tab_bg")
        self.tabs.add("🎯  Мониторинг")
        self.tabs.add("📋  Очередь")
        self.tabs.add("📊  Статистика")
        self.tabs.add("🔧  Настройки")
        self._build_monitor_tab(self.tabs.tab("🎯  Мониторинг"))
        self._build_queue_tab(self.tabs.tab("📋  Очередь"))
        self._build_stats_tab(self.tabs.tab("📊  Статистика"))
        self._build_settings_tab(self.tabs.tab("🔧  Настройки"))

    def _build_statusbar(self):
        self.status_bar = ctk.CTkLabel(
            self, text="  ●  Готов к работе",
            font=("Segoe UI", self._fs("label"), "bold"),
            text_color=self._t("status_ok"),
            fg_color=self._t("status_bg"), anchor="w")
        self.status_bar.grid(row=2, column=0, sticky="ew", ipady=6)

    # ── Вкладка «Мониторинг» ──────────────────────────────────────────────────
    def _build_monitor_tab(self, parent):
        parent.grid_columnconfigure(0, weight=1)

        self._section_label(parent, "🔗  ССЫЛКА НА ТОВАР", 0)
        self.entry_url = ctk.CTkEntry(
            parent, height=44,
            placeholder_text="https://market.yandex.ru/product--название-товара/...",
            font=self._f("input", bold=True),
            fg_color=self._t("input_bg"),
            text_color=self._t("entry_text"),
            border_color=self._t("accent"),
            border_width=2)
        self.entry_url.grid(row=1, column=0, padx=20, pady=(4, 6), sticky="ew")
        self._register(self.entry_url, "input_bg")
        self._register(self.entry_url, "accent", "border_color")

        paste_btn = ctk.CTkButton(
            parent, text="📋  Вставить из буфера",
            height=34, font=self._f("btn", bold=True),
            fg_color=self._t("btn_paste_bg"), hover_color=self._t("btn_paste_hover"),
            command=self._paste_url)
        paste_btn.grid(row=2, column=0, sticky="w", padx=20, pady=(0, 14))
        self._register(paste_btn, "btn_paste_bg")

        row_frame = ctk.CTkFrame(parent, fg_color="transparent")
        row_frame.grid(row=3, column=0, sticky="ew", padx=20, pady=(0, 12))
        row_frame.grid_columnconfigure(0, weight=1)
        row_frame.grid_columnconfigure(1, weight=1)

        pc = ctk.CTkFrame(row_frame, fg_color=self._t("card_bg"), corner_radius=12)
        pc.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        self._register(pc, "card_bg")

        pt = ctk.CTkLabel(pc, text="💰  ЦЕЛЕВАЯ ЦЕНА (₽)",
                           font=self._f("label", bold=True), text_color=self._t("accent"))
        pt.pack(anchor="w", padx=14, pady=(14, 4))
        self._register(pt, "accent", "text_color")

        self.entry_price = ctk.CTkEntry(
            pc, height=44,
            placeholder_text="например: 5990",
            font=("Segoe UI", self._fs("input_price"), "bold"),
            fg_color=self._t("input_bg"),
            text_color=self._t("entry_text"),
            border_color=self._t("accent"),
            border_width=2)
        self.entry_price.pack(padx=14, pady=(0, 14), fill="x")
        self._register(self.entry_price, "input_bg")
        self._register(self.entry_price, "accent", "border_color")

        ic = ctk.CTkFrame(row_frame, fg_color=self._t("card_bg"), corner_radius=12)
        ic.grid(row=0, column=1, sticky="nsew", padx=(8, 0))
        self._register(ic, "card_bg")

        it = ctk.CTkLabel(ic, text="⏱  ИНТЕРВАЛ ПРОВЕРКИ",
                           font=self._f("label", bold=True), text_color=self._t("accent"))
        it.pack(anchor="w", padx=14, pady=(14, 4))
        self._register(it, "accent", "text_color")

        self.interval_var = ctk.StringVar(value="1 час")
        self.interval_menu = ctk.CTkOptionMenu(
            ic, values=list(INTERVAL_OPTIONS.keys()),
            variable=self.interval_var,
            font=self._f("btn"), height=44)
        self.interval_menu.pack(padx=14, pady=(0, 14), fill="x")

        self.add_btn = ctk.CTkButton(
            parent, text="＋  ДОБАВИТЬ В ОЧЕРЕДЬ",
            height=48, font=self._f("btn_big", bold=True),
            fg_color=self._t("btn_add_bg"), hover_color=self._t("btn_add_hover"),
            command=self._add_to_queue)
        self.add_btn.grid(row=4, column=0, padx=20, pady=(0, 8), sticky="ew")
        self._register(self.add_btn, "btn_add_bg")

        self.toggle_btn = ctk.CTkButton(
            parent, text="▶   ЗАПУСТИТЬ МОНИТОРИНГ",
            height=58, font=self._f("btn_big", bold=True),
            fg_color=self._t("btn_start_bg"), hover_color=self._t("btn_start_hover"),
            command=self._toggle_monitoring)
        self.toggle_btn.grid(row=5, column=0, padx=20, pady=(0, 16), sticky="ew")

        self._section_label(parent, "📄  ЛОГ СОБЫТИЙ", 6)
        self.log_box = ctk.CTkTextbox(
            parent, height=195,
            font=("Consolas", self._fs("log"), "bold"),
            fg_color=self._t("log_bg"),
            text_color=self._t("log_text"))
        self.log_box.grid(row=7, column=0, padx=20, pady=(4, 8), sticky="ew")
        self._log("Система инициализирована. Готова к работе.")
        if not PLAYWRIGHT_OK:
            self._log("⚠️  Playwright не установлен! pip install playwright && playwright install chromium")
        if not OPENPYXL_OK:
            self._log("⚠️  openpyxl не установлен! pip install openpyxl")

        excel_btn = ctk.CTkButton(
            parent, text="📊  Открыть Excel-журнал",
            height=36, font=self._f("btn", bold=True),
            fg_color=self._t("btn_excel_bg"), hover_color=self._t("btn_excel_hover"),
            command=self._open_excel)
        excel_btn.grid(row=8, column=0, padx=20, pady=(0, 14), sticky="w")
        self._register(excel_btn, "btn_excel_bg")

    def _section_label(self, parent, text, row):
        lbl = ctk.CTkLabel(parent, text=text,
                           font=self._f("label", bold=True), text_color=self._t("section_label"))
        lbl.grid(row=row, column=0, sticky="w", padx=22, pady=(10, 0))
        self._register(lbl, "section_label", "text_color")

    # ── Вкладка «Очередь» ────────────────────────────────────────────────────
    def _build_queue_tab(self, parent):
        parent.grid_columnconfigure(0, weight=1)
        parent.grid_rowconfigure(1, weight=1)

        hl = ctk.CTkLabel(parent, text="Товары в очереди мониторинга",
                           font=self._f("btn_big", bold=True), text_color=self._t("accent"))
        hl.grid(row=0, column=0, padx=20, pady=(14, 6), sticky="w")
        self._register(hl, "accent", "text_color")

        self.queue_frame = ctk.CTkScrollableFrame(parent, fg_color=self._t("card_bg2"))
        self.queue_frame.grid(row=1, column=0, sticky="nsew", padx=20, pady=(0, 14))
        self.queue_frame.grid_columnconfigure(0, weight=1)
        self._register(self.queue_frame, "card_bg2")
        self._refresh_queue_display()

    # ── Вкладка «Статистика» ─────────────────────────────────────────────────
    def _build_stats_tab(self, parent):
        parent.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(parent, text="Статистика сессии",
                     font=self._f("btn_big", bold=True), text_color=self._t("accent")).pack(pady=18)
        sc = ctk.CTkFrame(parent, fg_color=self._t("card_bg"), corner_radius=14)
        sc.pack(fill="x", padx=20, pady=4)
        self._register(sc, "card_bg")
        self.stat_checks = self._stat_row(sc, "Всего проверок",    "0", "#44AAFF")
        self.stat_found  = self._stat_row(sc, "Цена найдена",      "0", "#44FFAA")
        self.stat_deals  = self._stat_row(sc, "Скидок обнаружено", "0", "#FF4466")
        self.stat_errors = self._stat_row(sc, "Ошибок",            "0", "#FF8844")
        pl = ctk.CTkLabel(parent, text=f"📁  Журнал: {EXCEL_PATH}",
                          font=("Segoe UI", self._fs("label_sm")), text_color=self._t("label_dim"), wraplength=540)
        pl.pack(pady=14, padx=20)
        self._register(pl, "label_dim", "text_color")

    def _stat_row(self, parent, label, value, color):
        f = ctk.CTkFrame(parent, fg_color="transparent")
        f.pack(fill="x", padx=20, pady=8)
        ctk.CTkLabel(f, text=label, font=self._f("stat_label", bold=True),
                     text_color=self._t("label_sub"), width=210, anchor="w").pack(side="left")
        vl = ctk.CTkLabel(f, text=value, font=("Segoe UI", self._fs("stat_big"), "bold"),
                          text_color=color, width=80, anchor="e")
        vl.pack(side="right")
        return vl

    # ── Вкладка «Настройки» ──────────────────────────────────────────────────
    def _build_settings_tab(self, parent):
        parent.grid_columnconfigure(0, weight=1)

        # ── Блок: темы для слабовидящих ──────────────────────────────────────
        sec1 = ctk.CTkFrame(parent, fg_color=self._t("card_bg"), corner_radius=14)
        sec1.pack(fill="x", padx=20, pady=(14, 8))
        self._register(sec1, "card_bg")

        ctk.CTkLabel(sec1, text="👁  ТЕМЫ ДЛЯ СЛАБОВИДЯЩИХ",
                     font=self._f("label", bold=True), text_color=self._t("accent")
                     ).pack(anchor="w", padx=16, pady=(14, 6))
        self._register_themed_lbl(lambda p=sec1: ctk.CTkLabel(
            p, text="Выберите режим отображения, наиболее подходящий для вас:",
            font=("Segoe UI", self._fs("label_sm")), text_color=self._t("label_dim"),
            wraplength=560), sec1).pack(anchor="w", padx=16, pady=(0, 10))

        themes_grid = ctk.CTkFrame(sec1, fg_color="transparent")
        themes_grid.pack(fill="x", padx=16, pady=(0, 16))
        themes_grid.grid_columnconfigure(0, weight=1)
        themes_grid.grid_columnconfigure(1, weight=1)

        self._theme_btns = {}
        for i, (key, label) in enumerate(ACCESSIBILITY_LABELS.items()):
            col = i % 2
            row = i // 2
            is_active = (self._current_theme == key)
            btn = ctk.CTkButton(
                themes_grid, text=label,
                height=44,
                font=self._f("btn", bold=True),
                fg_color=self._t("btn_add_bg") if not is_active else self._t("accent"),
                hover_color=self._t("btn_add_hover") if not is_active else self._t("accent_hover"),
                text_color=self._t("entry_text") if is_active else self._t("label_main"),
                command=lambda k=key: self._on_theme_change(k))
            btn.grid(row=row, column=col, sticky="ew", padx=4, pady=4)
            self._theme_btns[key] = btn

        # ── Блок: размер шрифта ───────────────────────────────────────────────
        sec2 = ctk.CTkFrame(parent, fg_color=self._t("card_bg"), corner_radius=14)
        sec2.pack(fill="x", padx=20, pady=(0, 8))
        self._register(sec2, "card_bg")

        ctk.CTkLabel(sec2, text="🔤  РАЗМЕР ШРИФТА",
                     font=self._f("label", bold=True), text_color=self._t("accent")
                     ).pack(anchor="w", padx=16, pady=(14, 6))

        font_row = ctk.CTkFrame(sec2, fg_color="transparent")
        font_row.pack(fill="x", padx=16, pady=(0, 16))
        font_row.grid_columnconfigure(0, weight=1)
        font_row.grid_columnconfigure(1, weight=1)

        self._font_btns = {}
        for j, (mode, label) in enumerate([("default", "Обычный"), ("large", "Крупный шрифт")]):
            is_active = (self._font_mode == mode)
            btn = ctk.CTkButton(
                font_row, text=label,
                height=44,
                font=self._f("btn", bold=True),
                fg_color=self._t("btn_excel_bg") if not is_active else self._t("accent"),
                hover_color=self._t("btn_excel_hover") if not is_active else self._t("accent_hover"),
                text_color=self._t("entry_text") if is_active else self._t("label_main"),
                command=lambda m=mode: self._on_font_change(m))
            btn.grid(row=0, column=j, sticky="ew", padx=4, pady=4)
            self._font_btns[mode] = btn

        # ── Блок: уведомления ─────────────────────────────────────────────────
        sec3 = ctk.CTkFrame(parent, fg_color=self._t("card_bg"), corner_radius=14)
        sec3.pack(fill="x", padx=20, pady=(0, 14))
        self._register(sec3, "card_bg")

        ctk.CTkLabel(sec3, text="🔔  УВЕДОМЛЕНИЯ",
                     font=self._f("label", bold=True), text_color=self._t("accent")
                     ).pack(anchor="w", padx=16, pady=(14, 8))

        ctk.CTkButton(sec3, text="Отправить тестовое уведомление",
                       height=44, font=self._f("btn", bold=True),
                       fg_color=self._t("btn_excel_bg"),
                       hover_color=self._t("btn_excel_hover"),
                       command=lambda: send_os_notification("PriceWatcher", "Тест работает! ✅")
                       ).pack(fill="x", padx=16, pady=(0, 16))

    def _register_themed_lbl(self, factory, parent):
        lbl = factory(parent)
        self._register(lbl, "label_dim", "text_color")
        return lbl

    def _on_theme_change(self, key):
        """Обработчик смены темы доступности."""
        self._current_theme = key
        # Обновляем вид кнопок
        for k, btn in self._theme_btns.items():
            is_active = (k == key)
            btn.configure(
                fg_color=self._t("accent") if is_active else self._t("btn_add_bg"),
                hover_color=self._t("accent_hover") if is_active else self._t("btn_add_hover"),
                text_color=self._t("entry_text") if is_active else self._t("label_main"),
            )
        self._apply_theme()

    def _on_font_change(self, mode):
        """Обработчик смены размера шрифта."""
        self._font_mode = mode
        for m, btn in self._font_btns.items():
            is_active = (m == mode)
            btn.configure(
                fg_color=self._t("accent") if is_active else self._t("btn_excel_bg"),
                hover_color=self._t("accent_hover") if is_active else self._t("btn_excel_hover"),
                text_color=self._t("entry_text") if is_active else self._t("label_main"),
            )
        self._rebuild_tabs()

    # ── Старое окно настроек (для совместимости) ──────────────────────────────
    def _open_settings(self):
        # Переключаемся на вкладку настроек вместо отдельного окна
        try:
            self.tabs.set("🔧  Настройки")
        except Exception:
            # Если вкладка не найдена — открываем отдельное окно
            win = ctk.CTkToplevel(self)
            win.title("Настройки")
            win.geometry("460x420")
            win.resizable(False, False)
            win.grab_set()
            win.focus_force()
            win.grid_columnconfigure(0, weight=1)

            ctk.CTkLabel(win, text="⚙  Настройки",
                         font=self._f("btn_big", bold=True),
                         text_color=self._t("accent")).pack(pady=18)

            ctk.CTkButton(win, text="🔔  Тест уведомления", height=44,
                          font=self._f("btn", bold=True),
                          fg_color=self._t("btn_excel_bg"),
                          hover_color=self._t("btn_excel_hover"),
                          command=lambda: send_os_notification("PriceWatcher", "Тест работает! ✅")
                          ).pack(pady=12, padx=20, fill="x")

            ctk.CTkButton(win, text="👁  Темы для слабовидящих",
                          height=44, font=self._f("btn", bold=True),
                          fg_color=self._t("btn_paste_bg"),
                          hover_color=self._t("btn_paste_hover"),
                          command=lambda: self.tabs.set("🔧  Настройки")
                          ).pack(pady=12, padx=20, fill="x")

            ctk.CTkButton(win, text="🖼  Выбрать обои окна",
                          height=44, font=self._f("btn", bold=True),
                          fg_color=self._t("btn_add_bg"),
                          hover_color=self._t("btn_add_hover"),
                          command=self._choose_bg
                          ).pack(pady=12, padx=20, fill="x")

            ctk.CTkButton(win, text="✕  Сбросить обои",
                          height=36, font=self._f("btn"),
                          fg_color=self._t("btn_stop_bg"),
                          hover_color=self._t("btn_stop_hover"),
                          command=self._clear_bg
                          ).pack(pady=12, padx=20, fill="x")

    def _choose_bg(self):
        path = filedialog.askopenfilename(
            title="Выберите изображение фона",
            filetypes=[("Images", "*.png *.jpg *.jpeg *.gif *.bmp"), ("All", "*.*")])
        if path: self._set_bg(path)

    def _set_bg(self, path):
        try:
            img = Image.open(path).resize((self.winfo_width() or 660, self.winfo_height() or 840))
            self._bg_image = ImageTk.PhotoImage(img)
            bg_label = tk.Label(self, image=self._bg_image)
            bg_label.place(relwidth=1, relheight=1)
            bg_label.lower()
            self.bg_path = path
        except ImportError:
            messagebox.showinfo("Нужен Pillow", "Установите: pip install Pillow")
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    def _clear_bg(self):
        for w in self.winfo_children():
            if isinstance(w, tk.Label): w.destroy()
        self.bg_path = None

    # ── Остальная логика (без изменений) ──────────────────────────────────────
    def _paste_url(self):
        try:
            t = self.clipboard_get()
            self.entry_url.delete(0, "end")
            self.entry_url.insert(0, t.strip())
        except Exception: pass

    def _add_to_queue(self):
        url = self.entry_url.get().strip()
        if not url:
            messagebox.showerror("Ошибка", "Введите ссылку на товар!"); return
        try:
            target = int(self.entry_price.get().strip())
        except ValueError:
            messagebox.showerror("Ошибка", "Введите целевую цену числом!"); return
        clean = clean_yandex_url(url)
        interval = self.interval_var.get()
        self.monitored_items.append({"url": clean, "target": target, "interval": interval})
        self._log(f"Добавлен: {clean[:55]}... | Цель: {target:,} ₽ | {interval}")
        self._refresh_queue_display()
        self.entry_url.delete(0, "end")
        self.entry_price.delete(0, "end")

    def _refresh_queue_display(self):
        for w in self.queue_frame.winfo_children(): w.destroy()
        if not self.monitored_items:
            ctk.CTkLabel(self.queue_frame,
                         text="Очередь пуста.\nДобавьте товары на вкладке «Мониторинг».",
                         font=self._f("btn"), text_color=self._t("label_dim")).pack(pady=40)
            return
        for i, item in enumerate(self.monitored_items):
            card = ctk.CTkFrame(self.queue_frame, fg_color=self._t("card_bg"), corner_radius=10)
            card.pack(fill="x", padx=8, pady=5)
            card.grid_columnconfigure(1, weight=1)
            self._register(card, "card_bg")

            ctk.CTkLabel(card, text=f" {i+1} ",
                         font=self._f("btn", bold=True),
                         fg_color=self._t("accent"), text_color="#FFFFFF",
                         corner_radius=6, width=30
                         ).grid(row=0, column=0, padx=(10, 8), pady=(10, 2), sticky="nw")

            url_s = item["url"][:62] + "..." if len(item["url"]) > 62 else item["url"]
            ctk.CTkLabel(card, text=url_s, font=("Segoe UI", self._fs("label_sm")),
                         text_color=self._t("queue_url"), wraplength=460, anchor="w"
                         ).grid(row=0, column=1, padx=(0, 8), pady=(10, 2), sticky="w")

            ctk.CTkLabel(card,
                         text=f"🎯  Цель: {item['target']:,} ₽     ⏱  {item['interval']}",
                         font=self._f("btn", bold=True), text_color=self._t("queue_price")
                         ).grid(row=1, column=1, padx=(0, 8), pady=(0, 10), sticky="w")

            idx = i
            ctk.CTkButton(card, text="✕", width=34, height=34,
                          font=("Segoe UI", self._fs("btn"), "bold"),
                          fg_color=self._t("btn_stop_bg"), hover_color=self._t("btn_stop_hover"),
                          command=lambda j=idx: self._remove_item(j)
                          ).grid(row=0, column=2, rowspan=2, padx=10)

    def _remove_item(self, idx):
        if 0 <= idx < len(self.monitored_items):
            r = self.monitored_items.pop(idx)
            self._log(f"Удалён: {r['url'][:50]}...")
            self._refresh_queue_display()

    def _toggle_monitoring(self):
        if self.is_monitoring:
            self.is_monitoring = False
            self.toggle_btn.configure(
                text="▶   ЗАПУСТИТЬ МОНИТОРИНГ",
                fg_color=self._t("btn_start_bg"), hover_color=self._t("btn_start_hover"))
            self._set_status("■  Мониторинг остановлен", "#FF8844")
            self._log("Мониторинг остановлен.")
        else:
            if not self.monitored_items:
                messagebox.showerror("Ошибка", "Добавьте хотя бы один товар в очередь!"); return
            if not PLAYWRIGHT_OK:
                messagebox.showerror("Ошибка",
                    "Playwright не установлен!\npip install playwright\nplaywright install chromium"); return
            self.is_monitoring = True
            self.toggle_btn.configure(
                text="⏹   ОСТАНОВИТЬ МОНИТОРИНГ",
                fg_color=self._t("btn_stop_bg"), hover_color=self._t("btn_stop_hover"))
            self._set_status("●  Мониторинг активен", "#44FF88")
            self._log(f"Запуск: {len(self.monitored_items)} товаров в очереди")
            self.monitor_thread = threading.Thread(target=self._monitor_loop, daemon=True)
            self.monitor_thread.start()

    def _monitor_loop(self):
        next_check = {i: 0 for i in range(len(self.monitored_items))}
        while self.is_monitoring:
            now = time.time()
            for i, item in enumerate(self.monitored_items):
                if not self.is_monitoring: return
                if now >= next_check.get(i, 0):
                    self._check_one(item)
                    next_check[i] = time.time() + INTERVAL_OPTIONS.get(item["interval"], 3600)
                    self._log(f"⏱  Следующая проверка: {item['interval']}")
            time.sleep(30)

    def _check_one(self, item):
        url, target = item["url"], item["target"]
        self._log(f"🔍  Проверяю: {url[:55]}...")
        self._session["checks"] += 1
        self._update_stats()
        try:
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=True, args=[
                    "--no-sandbox", "--disable-blink-features=AutomationControlled"])
                context = browser.new_context(
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
                    viewport={"width": 1280, "height": 800}, locale="ru-RU")
                context.add_init_script(
                    "Object.defineProperty(navigator,'webdriver',{get:()=>undefined});"
                    "window.chrome={runtime:{}};")
                page = context.new_page()
                try:
                    page.goto(url, wait_until="domcontentloaded", timeout=60000)
                    time.sleep(3)
                    price, name = parse_price_from_page(page)
                    if price is not None:
                        self._session["found"] += 1
                        self._log(f"✅  Цена: {price:,} ₽  |  {(name or '—')[:40]}")
                        if price <= target:
                            self._session["deals"] += 1
                            self._log(f"🔥  СКИДКА!  {price:,} ₽ ≤ {target:,} ₽")
                            send_os_notification("PriceWatcher — СКИДКА!!!",
                                f"{(name or 'Товар')[:40]}\nЦена: {price:,}₽ (цель: {target:,}₽)")
                            log_to_excel(url, name or "", price, target, "Скидка!!!")
                        else:
                            log_to_excel(url, name or "", price, target, f"💰 {price:,}₽")
                    else:
                        self._session["errors"] += 1
                        self._log(":< Цена не найдена (капча или новый селектор)")
                        log_to_excel(url, "", None, target, ":< Не найдена")
                except Exception as e:
                    self._session["errors"] += 1
                    self._log(f":< Ошибка: {str(e)[:70]}")
                    log_to_excel(url, "", None, target, f"❌ {str(e)[:30]}")
                finally:
                    browser.close()
        except Exception as e:
            self._session["errors"] += 1
            self._log(f"❌  Playwright: {str(e)[:70]}")
        self._update_stats()

    def _log(self, text):
        ts = datetime.now().strftime("%H:%M:%S")
        try:
            self.log_box.insert("end", f"[{ts}]  {text}\n")
            self.log_box.see("end")
        except Exception: pass

    def _set_status(self, text, color="#44CC88"):
        try: self.status_bar.configure(text=f"  {text}", text_color=color)
        except Exception: pass

    def _update_stats(self):
        s = self._session
        try:
            self.stat_checks.configure(text=str(s["checks"]))
            self.stat_found.configure(text=str(s["found"]))
            self.stat_deals.configure(text=str(s["deals"]))
            self.stat_errors.configure(text=str(s["errors"]))
        except Exception: pass

    def _open_excel(self):
        try:
            if platform.system() == "Windows": os.startfile(str(EXCEL_PATH))
            elif platform.system() == "Linux": subprocess.Popen(["xdg-open", str(EXCEL_PATH)])
            elif platform.system() == "Darwin": subprocess.Popen(["open", str(EXCEL_PATH)])
        except Exception: messagebox.showinfo("Путь", str(EXCEL_PATH))


if __name__ == "__main__":
    app = PriceWatcherApp()
    app.mainloop()
